#!/usr/bin/env python3
"""
Scrape Tamil Nadu assembly election results (May 2026) from ECI website.
Saves a CSV and a highlighted Excel workbook that colour-codes:
  - RED   : knife-edge fights (contest_index <= 10)
  - ORANGE: competitive but watchable (contest_index 10–40)
  - GREEN : virtually decided (projected margin >= 20,000 OR pct_complete >= 0.75
            with margin >= 5,000)

Usage:
    python3 scrape_election_results.py
"""

import subprocess
import sys

from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE_URL = "https://results.eci.gov.in/ResultAcGenMay2026/"
PAGE_URLS = [f"{BASE_URL}statewiseS22{i}.htm" for i in range(1, 13)]

COLUMNS = [
    "Constituency", "Const. No.", "Leading Candidate", "Leading Party",
    "Trailing Candidate", "Trailing Party", "Margin", "Round", "Status",
]

# Highlight fills
FILL_RED    = PatternFill("solid", fgColor="FF6B6B")   # knife-edge
FILL_ORANGE = PatternFill("solid", fgColor="FFB347")   # competitive
FILL_GREEN  = PatternFill("solid", fgColor="77DD77")   # decided
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # header row

FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(bold=True)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Fetch & parse
# ---------------------------------------------------------------------------

def fetch_html(url: str) -> str:
    result = subprocess.run(
        [
            "curl", "-s", "-L",
            "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "-H", "Accept-Language: en-US,en;q=0.9",
            url,
        ],
        capture_output=True, text=True, timeout=30,
    )
    if result.returncode != 0:
        raise RuntimeError(f"curl failed for {url}: {result.stderr}")
    return result.stdout


def cell_text(td) -> str:
    nested = td.find("table")
    if nested:
        first_td = nested.find("td")
        return first_td.get_text(strip=True) if first_td else ""
    return td.get_text(strip=True)


def parse_page(html: str) -> list[list]:
    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")
    if not tables:
        return []
    tbody = tables[0].find("tbody")
    if not tbody:
        return []
    rows = []
    for tr in tbody.children:
        if getattr(tr, "name", None) != "tr":
            continue
        cells = [cell_text(td) for td in tr.find_all(["td", "th"], recursive=False)]
        if not cells or all(c == "" for c in cells):
            continue
        cells = cells[:9] + [""] * (9 - len(cells))
        rows.append(cells)
    return rows


def scrape_all() -> pd.DataFrame:
    all_rows = []
    for i, url in enumerate(PAGE_URLS, start=1):
        print(f"  Fetching page {i:2d}/12: {url}")
        try:
            html = fetch_html(url)
            rows = parse_page(html)
            print(f"             → {len(rows)} rows")
            all_rows.extend(rows)
        except Exception as exc:
            print(f"             → ERROR: {exc}", file=sys.stderr)
    return pd.DataFrame(all_rows, columns=COLUMNS)


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def add_scores(df: pd.DataFrame) -> pd.DataFrame:
    split = df["Round"].astype(str).str.extract(r"(\d+)/(\d+)").astype(float)
    df["rounds_done"]  = split[0]
    df["rounds_total"] = split[1]
    df["pct_complete"] = df["rounds_done"] / df["rounds_total"]

    df["projected_margin"] = (
        (df["Margin"] / df["rounds_done"]) * df["rounds_total"]
    ).round(0)

    # contest_index: lower = tighter fight
    df["contest_index"] = df["Margin"] / (df["pct_complete"] * df["rounds_total"])

    return df


def classify(row) -> str:
    ci  = row["contest_index"]
    pm  = row["projected_margin"]
    pct = row["pct_complete"]
    m   = row["Margin"]

    # Knife-edge: tiny contest_index OR projected margin under 500
    if ci <= 10 or pm < 500:
        return "knife-edge"
    # Competitive: contest_index in range OR projected margin under 3000
    if ci <= 40 or pm < 3_000:
        return "competitive"
    if pm >= 20_000 or (pct >= 0.75 and m >= 5_000):
        return "decided"
    return ""


# ---------------------------------------------------------------------------
# Clean
# ---------------------------------------------------------------------------

def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include="str").columns:
        df[col] = df[col].str.strip()
    for col in ("Margin", "Const. No."):
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(",", ""), errors="coerce"
            )
    return df


# ---------------------------------------------------------------------------
# Excel output with highlighting
# ---------------------------------------------------------------------------

COL_WIDTHS = {
    "Constituency": 22, "Const. No.": 10, "Leading Candidate": 28,
    "Leading Party": 38, "Trailing Candidate": 28, "Trailing Party": 38,
    "Margin": 10, "Round": 8, "Status": 22,
}

LEGEND = [
    ("knife-edge", "FF6B6B", "Contest index ≤ 10 OR projected margin < 500 — could flip any round"),
    ("competitive", "FFB347", "Contest index 10–40 OR projected margin < 3000 — still very much in play"),
    ("decided",     "77DD77", "Projected margin ≥ 20k or counting 75%+ done with big lead"),
]


def write_excel(df: pd.DataFrame, path: str) -> None:
    # Write base data (score columns excluded from sheet)
    export_cols = COLUMNS + ["projected_margin", "contest_index"]
    df[export_cols].to_excel(path, index=False, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Results"

    # Column widths & header style
    for col_idx, col_name in enumerate(export_cols, start=1):
        letter = get_column_letter(col_idx)
        width = COL_WIDTHS.get(col_name, 16)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Map column name → index for lookup
    col_map = {name: idx + 1 for idx, name in enumerate(export_cols)}
    cat_col = col_map.get("contest_index")  # last extra column

    # Classify column (not written, used for fill)
    category_by_row = {}
    for df_idx, row in df.iterrows():
        category_by_row[df_idx] = classify(row)

    fill_map = {
        "knife-edge": FILL_RED,
        "competitive": FILL_ORANGE,
        "decided":     FILL_GREEN,
    }

    for ws_row, (df_idx, _) in enumerate(df.iterrows(), start=2):
        cat = category_by_row[df_idx]
        fill = fill_map.get(cat)
        for col_idx in range(1, len(export_cols) + 1):
            cell = ws.cell(row=ws_row, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill
            if col_idx == col_map["Constituency"] and cat:
                cell.font = FONT_BOLD

    # Number formats
    for ws_row in range(2, ws.max_row + 1):
        m_cell = ws.cell(row=ws_row, column=col_map["Margin"])
        m_cell.number_format = "#,##0"
        pm_cell = ws.cell(row=ws_row, column=col_map["projected_margin"])
        pm_cell.number_format = "#,##0"
        ci_cell = ws.cell(row=ws_row, column=col_map["contest_index"])
        ci_cell.number_format = "0.0"

    # --- Legend sheet ---
    leg = wb.create_sheet("Legend")
    leg.column_dimensions["A"].width = 14
    leg.column_dimensions["B"].width = 12
    leg.column_dimensions["C"].width = 60

    leg["A1"], leg["B1"], leg["C1"] = "Category", "Colour", "Meaning"
    for cell in [leg["A1"], leg["B1"], leg["C1"]]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    for r, (label, hex_color, desc) in enumerate(LEGEND, start=2):
        leg.cell(row=r, column=1, value=label).font = FONT_BOLD
        color_cell = leg.cell(row=r, column=2)
        color_cell.fill = PatternFill("solid", fgColor=hex_color)
        leg.cell(row=r, column=3, value=desc)

    leg["A6"] = "contest_index"
    leg["C6"] = "= current margin ÷ (fraction of rounds done × total rounds). Lower = tighter."
    leg["A7"] = "projected_margin"
    leg["C7"] = "= (margin ÷ rounds done) × total rounds. Assumes uniform vote distribution per round."

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Scraping 12 pages from {BASE_URL}\n")
    df = scrape_all()
    df = clean_df(df)
    df = add_scores(df)

    total = len(df)
    knife  = (df["contest_index"] <= 10).sum()
    comp   = ((df["contest_index"] > 10) & (df["contest_index"] <= 40)).sum()
    dec    = df.apply(classify, axis=1).eq("decided").sum()

    print(f"\nTotal constituencies : {total}")
    print(f"Knife-edge (red)     : {knife}")
    print(f"Competitive (orange) : {comp}")
    print(f"Virtually decided (green): {dec}")

    csv_path  = "tn_election_results_may2026.csv"
    xlsx_path = "tn_election_results_may2026.xlsx"

    df[COLUMNS].to_csv(csv_path, index=False)
    print(f"\nCSV   saved → {csv_path}")

    write_excel(df, xlsx_path)
    print(f"Excel saved → {xlsx_path}  (with highlights + Legend sheet)")


if __name__ == "__main__":
    main()
